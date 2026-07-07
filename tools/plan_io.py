#!/usr/bin/env python3
"""Read/write helpers for content_plan.xlsx so the ideation funnel never
fumbles openpyxl. Sheets: Plan (article ledger), Desire Library (eternal
territories), Seasonal Calendar (windows). This is the agents' I/O to the plan.
"""
from pathlib import Path
from datetime import date
from openpyxl import load_workbook

WB = Path(__file__).resolve().parent.parent / "content_plan.xlsx"

PLAN_COLS = ["Status", "Title / angle", "Type", "Theme", "Listings",
             "Scene refs (FILES in refs/scenes/)",
             "Infographic ref (refs/infographics/)", "Notes", "Date", "Slug"]


def _rows(ws):
    headers = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        out.append({h: (v if v is not None else "") for h, v in zip(headers, r)})
    return out


def sheet_rows(name):
    wb = load_workbook(WB, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        return []
    return _rows(wb[name])


def desire_library():
    return [r for r in sheet_rows("Desire Library")
            if r.get("Territory (desire / identity / aesthetic)")
            and "EXAMPLE" not in str(r.get("Territory (desire / identity / aesthetic)"))]


def seasonal_calendar():
    return sheet_rows("Seasonal Calendar")


def existing_titles():
    """For dedup — every title/slug already in the plan (any status)."""
    seen = set()
    for r in sheet_rows("Plan"):
        for k in ("Title / angle", "Slug"):
            v = str(r.get(k, "")).strip().lower()
            if v:
                seen.add(v)
    return seen


def append_plan_rows(rows, status="idea"):
    """rows: list of dicts using PLAN_COLS keys (missing keys -> blank).
    Skips any whose title already exists (dedup)."""
    wb = load_workbook(WB)
    ws = wb["Plan"]
    seen = existing_titles()
    added = 0
    for r in rows:
        title = str(r.get("Title / angle", "")).strip()
        if not title or title.lower() in seen:
            continue
        ws.append([r.get("Status", status)] + [r.get(c, "") for c in PLAN_COLS[1:]])
        seen.add(title.lower())
        added += 1
    wb.save(WB)
    return added


def mark_published(slug, when=None):
    wb = load_workbook(WB)
    ws = wb["Plan"]
    hit = False
    for row in ws.iter_rows(min_row=2):
        cells = {c.value: row[i] for i, c in enumerate(ws[1])}
        if str(cells["Slug"].value).strip() == slug.strip():
            cells["Status"].value = "published"
            cells["Date"].value = (when or date.today().isoformat())
            hit = True
    wb.save(WB)
    return hit


if __name__ == "__main__":
    import json, sys
    what = sys.argv[1] if len(sys.argv) > 1 else "desire"
    print(json.dumps({"desire": desire_library, "calendar": seasonal_calendar,
                      "plan": lambda: sheet_rows("Plan")}[what](),
                     indent=2, ensure_ascii=False))
